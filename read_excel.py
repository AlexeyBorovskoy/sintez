#!/usr/bin/env python3
"""
Скрипт для чтения Excel файла с протоколами SNMP
"""

import sys
import os

excel_file = "/home/alexey/shared_vm/Sintez_snmp_protocols(1).xlsx"

# Попробуем использовать openpyxl
try:
    import openpyxl
    print("Используется openpyxl")
    
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    print(f"\n📊 Файл: {os.path.basename(excel_file)}")
    print(f"📋 Количество листов: {len(wb.sheetnames)}")
    print(f"\n📑 Список листов:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {sheet_name}")
    
    print("\n" + "="*80)
    
    # Читаем каждый лист
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n📄 Лист: {sheet_name}")
        print(f"   Строк: {sheet.max_row}, Столбцов: {sheet.max_column}")
        
        # Показываем первые несколько строк
        print("\n   Первые строки:")
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_idx > 10:  # Показываем только первые 10 строк
                break
            row_data = [str(cell) if cell is not None else "" for cell in row]
            if any(row_data):  # Пропускаем пустые строки
                print(f"   {row_idx}: {' | '.join(row_data[:5])}")  # Первые 5 столбцов
        
        print()
    
except ImportError:
    # Попробуем pandas
    try:
        import pandas as pd
        print("Используется pandas")
        
        # Читаем все листы
        excel_file_obj = pd.ExcelFile(excel_file)
        
        print(f"\n📊 Файл: {os.path.basename(excel_file)}")
        print(f"📋 Количество листов: {len(excel_file_obj.sheet_names)}")
        print(f"\n📑 Список листов:")
        for i, sheet_name in enumerate(excel_file_obj.sheet_names, 1):
            print(f"  {i}. {sheet_name}")
        
        print("\n" + "="*80)
        
        # Читаем каждый лист
        for sheet_name in excel_file_obj.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            print(f"\n📄 Лист: {sheet_name}")
            print(f"   Размер: {df.shape[0]} строк × {df.shape[1]} столбцов")
            
            # Показываем первые строки
            print("\n   Первые строки:")
            print(df.head(10).to_string(max_cols=5))
            print()
            
    except ImportError:
        print("Ошибка: Не найдены библиотеки для чтения Excel файлов")
        print("\nУстановите одну из библиотек:")
        print("  pip install openpyxl")
        print("  или")
        print("  pip install pandas openpyxl")
        sys.exit(1)
    except Exception as e:
        print(f"Ошибка при чтении файла: {e}")
        sys.exit(1)
except Exception as e:
    print(f"Ошибка при чтении файла: {e}")
    sys.exit(1)
